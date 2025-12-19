# ==============================================
# DIŞA AKTARMA (EXPORT) MODÜLÜ
# ==============================================
# Bu dosya sınav programını PDF ve Excel
# formatlarında dışa aktarma işlemlerini yapar.
# Takvim Formatı: Tarihler sütun, saatler satır
# Her bölüm için ayrı tablo
# ==============================================

import os
from datetime import datetime
from collections import defaultdict

# PDF oluşturma için
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel oluşturma için
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.models.exam import get_all_exams
from app.database import execute_query


def get_export_folder():
    """Export klasörünün yolunu döndürür."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_folder = os.path.join(base_dir, 'exports')
    if not os.path.exists(export_folder):
        os.makedirs(export_folder)
    return export_folder


def get_departments():
    """Tüm bölümleri getirir."""
    query = "SELECT id, name FROM departments ORDER BY name"
    return execute_query(query)


def get_exams_by_department(department_id):
    """Bölüme göre sınavları getirir."""
    query = """
        SELECT e.*, 
               c.code as course_code, 
               c.name as course_name,
               c.student_count,
               cl.name as classroom_name,
               s.name as supervisor_name,
               i.name as instructor_name
        FROM exam_schedule e
        LEFT JOIN courses c ON e.course_id = c.id
        LEFT JOIN classrooms cl ON e.classroom_id = cl.id
        LEFT JOIN instructors s ON e.supervisor_id = s.id
        LEFT JOIN instructors i ON c.instructor_id = i.id
        WHERE c.department_id = ?
        ORDER BY e.exam_date, e.start_time
    """
    return execute_query(query, (department_id,))


def organize_exams_as_calendar(exams):
    """
    Sınavları takvim formatında organize eder.
    
    Döndürür:
        dates: Benzersiz tarihler listesi (DD.MM.YYYY formatında)
        time_slots: Benzersiz saat dilimleri listesi
        calendar_data: {(tarih, saat): exam_info} sözlüğü
    """
    # Tarihleri DD.MM.YYYY formatına çevir
    def format_date(date_str):
        if '-' in date_str and len(date_str) == 10:
            parts = date_str.split('-')
            return f"{parts[2]}.{parts[1]}.{parts[0]}"
        return date_str
    
    raw_dates = sorted(set(exam['exam_date'] for exam in exams))
    dates = [format_date(d) for d in raw_dates]
    time_slots = sorted(set(exam['start_time'] + '-' + exam['end_time'] for exam in exams))
    
    calendar_data = {}
    for exam in exams:
        date = format_date(exam['exam_date'])  # Formatlanmış tarih kullan
        time_slot = exam['start_time'] + '-' + exam['end_time']
        key = (date, time_slot)
        
        exam_info = f"{exam['course_code']}\n{exam['classroom_name']}"
        
        # supervisor_name veya instructor_name'i kontrol et
        supervisor = None
        try:
            supervisor = exam['supervisor_name']
        except (KeyError, IndexError):
            pass
        
        if not supervisor:
            try:
                supervisor = exam['instructor_name']
            except (KeyError, IndexError):
                pass
        
        if supervisor:
            exam_info += f"\n({supervisor})"
        
        if key in calendar_data:
            calendar_data[key] += '\n---\n' + exam_info
        else:
            calendar_data[key] = exam_info
    
    return dates, time_slots, calendar_data


def register_fonts():
    """Türkçe karakter desteği için font kaydet."""
    font_name = 'Helvetica'
    font_name_bold = 'Helvetica-Bold'
    font_registered = False
    
    # Windows Arial
    try:
        arial_path = 'C:/Windows/Fonts/arial.ttf'
        arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
        
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('Arial', arial_path))
            if os.path.exists(arial_bold_path):
                pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bold_path))
                font_name_bold = 'Arial-Bold'
            font_name = 'Arial'
            font_registered = True
    except:
        pass
    
    return font_name, font_name_bold, font_registered


def convert_turkish(text, font_registered):
    """Türkçe karakterleri ASCII'ye çevir (font yoksa)."""
    if font_registered or not text:
        return text
    
    tr_chars = {'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G', 
               'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S',
               'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'}
    for tr, en in tr_chars.items():
        text = text.replace(tr, en)
    return text


def export_to_pdf():
    """
    Sınav programını takvim formatında PDF olarak dışa aktarır.
    Her bölüm için ayrı tablo.
    """
    # Dosya oluştur
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = 'sinav_programi_' + timestamp + '.pdf'
    file_path = os.path.join(get_export_folder(), filename)
    
    # Font kaydet
    font_name, font_name_bold, font_registered = register_fonts()
    
    # PDF belgesi oluştur
    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=0.5*cm,
        leftMargin=0.5*cm,
        topMargin=0.5*cm,
        bottomMargin=0.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=16,
        alignment=1,
        spaceAfter=10
    )
    
    dept_style = ParagraphStyle(
        'DeptTitle',
        parent=styles['Heading2'],
        fontName=font_name_bold,
        fontSize=12,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=5,
        spaceBefore=10
    )
    
    elements = []
    
    # Ana başlık
    title = Paragraph("SINAV PROGRAMI", title_style)
    elements.append(title)
    
    today = datetime.now().strftime('%d.%m.%Y')
    subtitle = Paragraph(f"Olusturulma Tarihi: {today}", styles['Normal'])
    elements.append(subtitle)
    elements.append(Spacer(1, 0.5*cm))
    
    # Bölümleri al
    departments = get_departments()
    
    for dept in departments:
        dept_id = dept['id']
        dept_name = convert_turkish(dept['name'], font_registered)
        
        # Bu bölümün sınavlarını al
        exams = get_exams_by_department(dept_id)
        
        if len(exams) == 0:
            continue
        
        # Bölüm başlığı
        dept_title = Paragraph(f"📚 {dept_name}", dept_style)
        elements.append(dept_title)
        
        # Takvim formatına dönüştür
        dates, time_slots, calendar_data = organize_exams_as_calendar(exams)
        
        if not dates or not time_slots:
            continue
        
        # Tablo verisi oluştur
        # İlk satır: boş + tarihler
        header_row = ['Saat'] + dates
        table_data = [header_row]
        
        # Saat satırları
        for time_slot in time_slots:
            row = [time_slot]
            for date in dates:
                key = (date, time_slot)
                cell_value = calendar_data.get(key, '')
                cell_value = convert_turkish(cell_value, font_registered)
                row.append(cell_value)
            table_data.append(row)
        
        # Sütun genişlikleri
        num_cols = len(dates) + 1
        col_width = (26*cm) / num_cols
        col_widths = [col_width] * num_cols
        col_widths[0] = 2.5*cm  # Saat sütunu
        
        # Tablo oluştur
        table = Table(table_data, colWidths=col_widths)
        
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e5e7eb')),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTNAME', (0, 0), (0, -1), font_name_bold),
            ('FONTNAME', (1, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        elements.append(PageBreak())  # Her bölüm yeni sayfada
    
    # Sınav yoksa
    if len(elements) <= 3:
        elements.append(Paragraph("Henuz planlanmis sinav bulunmamaktadir.", styles['Normal']))
    
    doc.build(elements)
    return file_path


def export_to_excel():
    """
    Sınav programını takvim formatında Excel olarak dışa aktarır.
    Her bölüm için ayrı sayfa.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = 'sinav_programi_' + timestamp + '.xlsx'
    file_path = os.path.join(get_export_folder(), filename)
    
    workbook = Workbook()
    
    # İlk sayfayı sil
    default_sheet = workbook.active
    
    # Stiller
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    time_fill = PatternFill(start_color='e5e7eb', end_color='e5e7eb', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bölümleri al
    departments = get_departments()
    sheet_created = False
    
    for dept in departments:
        dept_id = dept['id']
        dept_name = dept['name'][:31]  # Excel sayfa adı limiti
        
        # Bu bölümün sınavlarını al
        exams = get_exams_by_department(dept_id)
        
        if len(exams) == 0:
            continue
        
        # Yeni sayfa oluştur
        if not sheet_created:
            sheet = default_sheet
            sheet.title = dept_name
            sheet_created = True
        else:
            sheet = workbook.create_sheet(title=dept_name)
        
        # Takvim formatına dönüştür
        dates, time_slots, calendar_data = organize_exams_as_calendar(exams)
        
        if not dates or not time_slots:
            continue
        
        # Başlık satırı
        sheet.cell(row=1, column=1, value='Saat')
        sheet.cell(row=1, column=1).font = header_font
        sheet.cell(row=1, column=1).fill = header_fill
        sheet.cell(row=1, column=1).alignment = cell_alignment
        sheet.cell(row=1, column=1).border = thin_border
        
        for col_num, date in enumerate(dates, 2):
            cell = sheet.cell(row=1, column=col_num, value=date)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # Saat satırları
        for row_num, time_slot in enumerate(time_slots, 2):
            # Saat sütunu
            time_cell = sheet.cell(row=row_num, column=1, value=time_slot)
            time_cell.font = Font(bold=True)
            time_cell.fill = time_fill
            time_cell.alignment = cell_alignment
            time_cell.border = thin_border
            
            # Tarih sütunları
            for col_num, date in enumerate(dates, 2):
                key = (date, time_slot)
                cell_value = calendar_data.get(key, '')
                
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Sınav varsa arka plan rengi
                if cell_value:
                    cell.fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
        
        # Sütun genişlikleri
        sheet.column_dimensions['A'].width = 12
        for col_num in range(2, len(dates) + 2):
            col_letter = chr(64 + col_num)
            sheet.column_dimensions[col_letter].width = 18
        
        # Satır yükseklikleri
        for row_num in range(1, len(time_slots) + 2):
            sheet.row_dimensions[row_num].height = 50
    
    # Hiç sayfa oluşturulmadıysa
    if not sheet_created:
        default_sheet.title = "Boş"
        default_sheet.cell(row=1, column=1, value="Henüz planlanmış sınav bulunmamaktadır.")
    
    workbook.save(file_path)
    return file_path
